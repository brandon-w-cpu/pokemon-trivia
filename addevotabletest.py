from openpyxl import load_workbook
import pymongo

workbook = load_workbook("All Evo.xlsx", data_only = True) 
sheet = workbook.active
myclient = pymongo.MongoClient("mongodb://localhost:27017/?readPreference=primary&appname=MongoDB%20Compass&ssl=false")
mydb = myclient["PokemonTrivia"]

mycol = mydb["LevelEvos"]
mycol.drop()
mycol = mydb["LevelEvos"]

for index, row in enumerate(sheet.iter_rows()): 
    if(index > 0):
        name = str(row[0].value).lower().strip()
        evolvesinto = str(row[2].value).lower().strip()
        requirement = str(row[3].value).lower().strip()
        specialrequirements = str(row[4].value).lower().strip()
        if(name != evolvesinto and ((name != "None") and (evolvesinto != "None"))):
            print(row[0].value)
            mydict = {"name": row[0].value, "evolvesinto": row[2].value, "category": "Level", "requirement": row[3].value, "specialrequirements" : row[4].value}
            x = mycol.insert_one(mydict)

workbook = load_workbook("Moves.xlsx", data_only = True) 
sheet = workbook.active
mycol = mydb["Moves"]
mycol.drop()
mycol = mydb["Moves"]

for index, row in enumerate(sheet.iter_rows()): 
    if(index > 0):
        name = str(row[0].value).lower().strip()
        typing = str(row[1].value).lower().strip()
        category = str(row[2].value).lower().strip()
        contest = str(row[3].value).lower().strip()
        pp = str(row[4].value).lower().strip()
        power = str(row[5].value).lower().strip()
        accuracy = str(row[6].value).lower().strip()
        gen = str(row[7].value).lower().strip()
        print(row[0].value)
        mydict = {"name": name, "typing": typing, "category": category, "contest" : contest, "pp" : pp, "power": power, "accuracy": accuracy, "gen" : gen}
        x = mycol.insert_one(mydict)